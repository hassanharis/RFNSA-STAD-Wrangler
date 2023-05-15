#!/usr/bin/env python
# coding: utf-8

# # RFNSA Antenna Details
# ##### ©Haris Hassan
# 

# In[27]:


################ Import Libraries ################

from IPython import get_ipython;   
get_ipython().magic('reset -sf')
import pandas as pd
pd.set_option('display.max_colwidth', None)
from IPython.display import display, HTML
import re
import math
import numpy as np
from itertools import count
class text_format:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'
    
# Function to convert from mW to dBm
def W2dBm(mW):
    return 10.*log10(mW) + 30


# Function to convert from dBm to mW
def dBm2W(dBm):
    return 10**((dBm-30)/10.)

################################################

    
################### Replace with path to your excel file. 
data = pd.read_excel(r'C:\Users\Mewtwo\Downloads\New Microsoft Excel Worksheet.xlsx')

#remove the proposed Antennas
data = data.drop(data[data['Existing/Proposed'] == 'Proposed'].index)

################ Create the dataframe from meaningful columns of STAD table and rename the columns ################
df = pd.DataFrame(data, columns=['Antenna ID No','Antenna Model','Sector','Height - Phase Centre (m)'
                                 ,'Bearing Degrees (true)',
                                 'Mech Downtilt','Elect Downtilt','System',
                                 'Port Number (Band Power per Port (dBm))','Band Power per Port (dBm)','Notes'])
df = df.rename(columns={'Antenna Model': 'Antenna', 'Antenna ID No': 'ID', 'Bearing Degrees (true)': 'Bearing'
                  , 'Mech Downtilt': 'MDT', 'Elect Downtilt': 'EDT', 'Height - Phase Centre (m)': 'Height', 
                   'Port Number (Band Power per Port (dBm))': 'Possible Ports', 
                   'Band Power per Port (dBm)': 'Powers'})

df.fillna(value = 0,inplace = True)

# Extracting Carrier (optus, telstra, Vodafone), technology/Frequency (i.e LTE900, NR2100, WCDMA850) information from System 
df['System'] = df['System'].str.replace(' \[', ';', regex=True)
df['System'] = df['System'].str.replace('/', ';', 1, regex=True)
df['Carrier'] = df['System'].str.split(" ;").str[0]
df['Tech'] = df['System'].str.split(";").str[1]

# Sorting the STAD dataframe to match with RFNSA
df.sort_values(by=['Carrier','ID'], ascending=[True, True], inplace = True)
#Reset index of dataframe after sorting and start from 1
df = df.reset_index(drop=True)
df.index += 1

#Count the total ports of each antenna
df['Total Ports'] = df['Powers'].str.count(';')
df.fillna(value = 0,inplace = True)
df['Total Ports'] = df['Total Ports'] + 1
df['Total Ports'] = list(x for x in df['Total Ports'])

#Format/Clean the Data
df['EDT'] = df['EDT'].str.replace(' to ', '-', regex=True)
df['EDT'] = df['EDT'].str.replace(r'\(.*\)','', regex=True)
df['Sector'] = df['Sector'].astype(str).replace('\.0', '', regex=True)
df['Sector']= [''.join(''.join(map(str, m))+' - Sector ' +f'{l}') for l, m in zip(df['Sector'],df['Tech'])]

#Correcting microwavelinks
templist = []
for x in df['Powers']:
    if ';' not in str(x):
        templist.append(str(x)+';')
    else:
        templist.append(str(x))        
df['Powers'] = templist


# In[28]:


### Convert te Power(dBm) into Watts
temp_power = []
for elm in df['Powers'].str.split(";"):
    if isinstance(elm, (list, tuple)):
        elm = [x for x in elm if x != '']  # Remove any empty strings from the list
        converted_values = []
        for sub in elm:
            if float(sub) >= 30:
                try:
                    converted_values.append(round(dBm2W(float(sub)),1))
                except ValueError:
                    converted_values.append(sub)
                    pass  # Skip the current value if it cannot be converted to a float
            else:
                 try:
                    converted_values.append(round(dBm2W(float(sub)),4))
                 except ValueError:
                    converted_values.append(sub)
                    pass  # Skip the current value if it cannot be converted to a float
        temp_power.append(converted_values)
    else:
        try:
            temp_power.append([dBm2W(float(elm))])
        except ValueError:
            temp_power.append([])

df['Powers (W)'] = temp_power


# In[29]:


display(df)


# In[30]:


######## what should be the frequency? ########
Assess_Freq_list = []
for i in df['Tech']:
    if any([x in i for x in ['GSM900', 'WCDMA850', 'NB-IOT900', 'WCDMA900','LTE850', 'NR850','LTE900']]):
        Assess_Freq_list.append('900')
    elif any([x in i for x in ['NR/LTE2100', 'WCDMA2100', 'LTE2100']]):
        Assess_Freq_list.append('2100')
    elif any([x in i for x in ['NR/LTE1800', 'LTE1800']]):
        Assess_Freq_list.append('1800')
    elif any([x in i for x in ['LTE2600', 'NR2600']]):
        Assess_Freq_list.append('2600')
    elif 'LTE700' in i:
        Assess_Freq_list.append('750')
    elif 'LTE2300' in i:
        Assess_Freq_list.append('2350')
    else:
        Assess_Freq_list.append(i)
df['Assess Freq'] = Assess_Freq_list


# In[31]:


######## Which ports to add powers to? ########
Mylist =[]

for i in df['Powers']:
    count = 1
    lister = []
    if isinstance(i, str):
        for j in i:
            if j==';':
                count+=1
            else:
                lister.append(count)
        Mylist.append(lister)
    else:
        Mylist.append(list(str(i))) 
        
    
new_k = []
for elem in Mylist:
    kiterator=[]
    for el in elem:
        if el not in kiterator:
            kiterator.append(el)
    new_k.append(kiterator)
k = new_k

df['WhereToAddPower'] = k


# In[32]:


####Formatting Powers Column with ports numbers

#itArr=[]
#for index, elm in df['Powers'].iteritems():
#    if isinstance(elm, (list, tuple)):
#        itArr.append([f'{i+1}. {l}' for i, l in enumerate(elm)])
#    else:
#        itArr.append([elm])
#df=df.drop(columns='Powers', axis=1)
#df.insert(2, 'Powers', itArr)
#############################################


# ## Add these Antennas to prox5

# In[33]:


AntennasUnique = list( dict.fromkeys(df['Antenna']) )
print('You need to add these antennas\n\n'+str.join(" \n", AntennasUnique)+' \n\nto PROX5')


# ## IDs for Antennas

# In[34]:


IDUnique = list( dict.fromkeys(df['ID'],df['Antenna']) )
AntennaIds = {}
for antenna in AntennasUnique:
    IdsForAntenna = list(dict.fromkeys(df.loc[df['Antenna'] == antenna, 'ID']))
    AntennaIds[antenna] = IdsForAntenna

for x, y in AntennaIds.items():
    print(f'\nThe id of '+text_format.BOLD+ x + text_format.END + ' are\n'+str.join(" \n", str(y)))


# ## Antenna Settings

# In[35]:


display('Set Power and Frequency to ports in Prox5')
Settings = {}
#for i in IDUnique:
#    IdsForAntenna = list(dict.fromkeys(df.loc[df['ID'] == i, 'ID']))
#    AntennaIds[antenna] = IdsForAntenna

for i in IDUnique:
        print(f'\nFor '+text_format.BOLD+str(i)+text_format.END+f', the Settings are')
        display(df.loc[(df['ID'] == i),['ID','Height','Bearing','MDT']])


# ## Set Power and Frequency to ports in Prox5

# In[36]:


display('Set Power and Frequency to ports in Prox5')
for i in IDUnique:
        print(text_format.BOLD+str(i))
        display(df.loc[(df['ID'] == i),['ID','EDT','Tech','Possible Ports','WhereToAddPower','Powers (W)']])


# ## EMEG Equipment List

# In[45]:


totalPortsdict = {}
for idantenna in IDUnique:
    totalPorts = list(dict.fromkeys(df.loc[df['ID'] == idantenna, 'Total Ports']))
    totalPortsdict[idantenna] = totalPorts
    
#display(totalPortsdict)   
#print(df.loc[(df['ID']=='11-O'), ['EDT','Sector','WhereToAddPower','Power (W)']])
#display(df.loc[(df['ID'] == i),['ID','Height','Bearing','MDT']])

listpow = []
listsec = []
x = 0
yprev = 0
for AntId, AntPorts in totalPortsdict.items():
    templistpow = [''] * (int(AntPorts[0] / 2) + int(AntPorts[0] % 2))
    templistsec = [''] * (int(AntPorts[0] / 2) + int(AntPorts[0] % 2))
    tempdf = df.loc[(df['ID']==AntId), ['EDT','Sector','WhereToAddPower','Powers (W)']]
    for sector, port, power in zip(tempdf['Sector'], tempdf['WhereToAddPower'], tempdf['Powers (W)']):
        display(AntId)
        if len(port)>1:
            for x, y in enumerate(port):       
                if sector not in templistsec[math.ceil(int(y)/2)-1]:
                    templistsec[math.ceil(int(y)/2)-1] += sector +'\n'
                    templistpow[math.ceil(int(y)/2)-1] += '\n'
                templistpow[math.ceil(int(y)/2)-1] += '+' + str(power[x])
        if len(port)==1:
            for x, y in enumerate(port):
                display(math.ceil(int(y)/2)-1)
                templistsec[math.ceil(int(y)/2)-1] += sector +'\n'
                if y%2==0:
                    templistpow[math.ceil(int(y)/2)-1] += r'0 ++ ' + str(power[x])
                else:
                    templistpow[math.ceil(int(y)/2)-1] += str(power[x]) + ' ++ 0'
            
    listpow.append(templistpow)
    listsec.append(templistsec)

for index1, x in enumerate(listpow):
    for index2, y in enumerate(x):
        if y == '':
            listpow[index1][index2] = '0'
        
for index1, x in enumerate(listsec):
    for index2, y in enumerate(x):
        if y == '':
            listsec[index1][index2] = '-'
        

EquipmentList = pd.DataFrame(totalPortsdict.items(), columns=['ID', 'No. of ports'])
EquipmentList['System/Sector'] = listsec
EquipmentList['Power (W)'] = listpow

EquipmentList_final = pd.DataFrame([], columns=['ID','System/Sector', 'Power (W)'])

for a, x, y in zip(EquipmentList['ID'],EquipmentList['System/Sector'], EquipmentList['Power (W)']): 
        EquipmentListAntenna = pd.DataFrame([a]*len(x),columns=['ID'])
        EquipmentListAntenna['System/Sector'] = x
        EquipmentListAntenna['Power (W)'] = y
        EquipmentList_final = pd.concat([EquipmentList_final, EquipmentListAntenna], axis=0)

EquipmentList_final['Power (W)'] = EquipmentList_final['Power (W)'].str.replace('\+', '', 1, regex=True)
EquipmentList_final['Power (W)'] = EquipmentList_final['Power (W)'].str.replace('\\n\+', '\\n', regex=True)
EquipmentList_final['Power (W)'] = EquipmentList_final['Power (W)'].str.strip()
EquipmentList_final['System/Sector'] = EquipmentList_final['System/Sector'].str.strip()
#display(EquipmentList_final)


# ## Exports

# In[46]:


#df3 = pd.DataFrame.from_dict(AntennaIds)
df3 = pd.DataFrame(dict([(k,pd.Series(v)) for k,v in AntennaIds.items() ]))
df3 = df3.fillna('')



dfAntennaSettings = df[['ID','Height','Bearing','MDT','Total Ports']].copy()
dfAntennaSettings=dfAntennaSettings.drop_duplicates(keep='first')

path = r'C:\Users\Mewtwo\Desktop'+'\\'+ data['Site Name'][3] + 'RFNSA Wrangled Data.xlsx'

from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
wb.save(path)

writer = pd.ExcelWriter(path, engine='openpyxl')
df3.to_excel(writer, sheet_name = 'Antenna IDs', index=False)
dfAntennaSettings.to_excel(writer, sheet_name = 'Antenna Settings', index=False)
df.to_excel(writer, sheet_name = 'Add powers to Prox', columns = ['Antenna','ID','EDT','Tech','Possible Ports','WhereToAddPower','Powers (W)','Assess Freq','Notes'], index=False)
EquipmentList_final.to_excel(writer, sheet_name = 'EMEG List', index=False)

# load the Excel file with openpyxl
workbook = writer.book

for sheet in workbook:
    for column in range(1, sheet.max_column + 2):
        letter = get_column_letter(column)
        #sheet.column_dimensions[letter].auto_size = True
        sheet.column_dimensions[letter].bestFit = True

# save the modified workbook
workbook.save(path)


# In[48]:


#display( HTML( df.to_html().replace("\\n","<br><li type='1'>") ) )
EquipmentList_final['Power (W)'] = EquipmentList_final['Power (W)'].str.replace(r'\n','<br>', regex=True)
#EquipmentList_final['System/Sector'] = EquipmentList_final['System/Sector'].str.replace((r'\n','<br>', regex=True)

EquipmentList_final['System/Sector'] = EquipmentList_final['System/Sector'].str.replace(r'\n','&#10;')

#EquipmentList_final.to_html(r'C:\Users\Mewtwo\Desktop\EMEGList.html', index=False)
#dfAntennaSettings.to_html(r'C:\Users\Mewtwo\Desktop\Antennadata.html', index=False)
#df3.to_html(r'C:\Users\Mewtwo\Desktop\Antennadata2.html',index=False)
#df.to_html(r'C:\Users\Mewtwo\Desktop\Antennadata3.html', columns = ['ID','EDT','Tech','Possible Ports','WhereToAddPower','Powers (W)','Assess Freq','Notes'], index=False)

#df.to_html(r'C:\Users\Mewtwo\Desktop\Antennadata4.html',columns = ['ID','EDT','Sector','WhereToAddPower','Powers (W)'],index=False)
#import subprocess
#subprocess.call('wkhtmltoimage -f png --width 0 Antennadata.html Antennadata.png', shell=True)


# In[49]:


for sheetname in workbook.sheetnames:
    worksheet = workbook[sheetname]

    # set best fit attribute to true for all columns
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        worksheet.column_dimensions[column_cells[0].column_letter].bestFit = True

worksheet = workbook.active

# set the width of column A to 15
column_letter = get_column_letter(1)  # A
column_dimension = worksheet.column_dimensions[column_letter]
column_dimension.width = 15
column_dimension.bestFit = True


# In[50]:


#with open("C:\\Users\\Mewtwo\\Downloads\\RFNSA Scrapper_V2.html", encoding="utf8") as html_file:
#    content = html_file.read()

# Get rid off prompts and source code
#content = content.replace("div.input_area {","div.input_area {\n\tdisplay: none;")    
#content = content.replace(".prompt {",".prompt {\n\tdisplay: none;")

#f = open(FILE, 'w')
#f.write(content)
#f.close()


# In[51]:


#jupyter nbconvert --to pdf 'RFNSA Scrapper_V2.ipynb' --template=hidecode.tplx


# In[ ]:




